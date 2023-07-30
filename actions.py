from flask import Blueprint, request, render_template, redirect, url_for, flash, Response, jsonify
import openpyxl
import warnings
import re

actions = Blueprint('actions', __name__, template_folder='templates')


@actions.route('/')
def index():
    return render_template('index.html')


@actions.route('/check-vietnamese', methods=['POST', 'GET'])
def check_vietnamese():
    try:
        if request.method == 'POST':
            print("# check_vietnamese #")
            checkData = []
            file = request.files['filevn']
            # Đường dẫn đến tệp Excel chứa test case
            # file_path = 'danh_sach.xlsx'

            # Tắt cảnh báo cho openpyxl
            warnings.simplefilter("ignore", category=UserWarning)

            # Mở tệp Excel
            workbook = openpyxl.load_workbook(file)

            # Lấy danh sách các tên SHEET trong FILE
            sheet_names = workbook.sheetnames

            # Duyệt qua từng trang tính trong tệp
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                number = 0
                # Duyệt qua từng ô trên trang tính
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for index, cell_value in enumerate(row, start=1):
                        if cell_value is not None and has_vietnamese(str(cell_value)):
                            number += 1
                            cell_address = openpyxl.utils.get_column_letter(index) + str(row_index)
                            checkDataErrorInfor = {
                                'index': number,
                                'sheet_name': sheet_name,
                                'cell_address': cell_address,
                                'cell_value': cell_value
                            }
                            checkData.append(checkDataErrorInfor)

            # Đóng tệp Excel
            workbook.close()
            print(len(checkData))
            return render_template('check-vietnamese.html', checkData=checkData)

        else:
            return render_template('check-vietnamese.html')

    except Exception as e:
        # Xảy ra lỗi
        print("error : ", e)
        return render_template('check-vietnamese.html')


def has_vietnamese(text):
    # Biểu thức chính quy để kiểm tra ký tự tiếng Việt
    pattern = r'[\u00C0-\u024F\u1E00-\u1EFF]'

    if re.search(pattern, text):
        return True
    else:
        return False


print(has_vietnamese("'Phân tích issue của Q1'"))


@actions.route('/check-japanese')
def check_japanese():
    return render_template('check-japanese.html')
